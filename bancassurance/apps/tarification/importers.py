"""
Importers Excel → Base de données
Permet d'importer les tables tarifaires depuis des fichiers Excel
"""
import openpyxl
from decimal import Decimal
from datetime import datetime, date
from django.db import transaction
from .models import (
    TableTauxEmprunteur,
    TableCIMA_H,
    TablePrimesEtudes,
    TableTauxMensuels,
    ParametresProduits
)


class ExcelImporter:
    """Classe de base pour les importers Excel"""
    
    def __init__(self, filepath):
        self.filepath = filepath
        self.workbook = None
        self.errors = []
        self.stats = {
            'created': 0,
            'updated': 0,
            'errors': 0
        }
    
    def load_workbook(self):
        """Charge le fichier Excel"""
        try:
            self.workbook = openpyxl.load_workbook(self.filepath, data_only=True)
            return True
        except Exception as e:
            self.errors.append(f"Erreur chargement fichier: {e}")
            return False
    
    def log_error(self, row, message):
        """Enregistre une erreur"""
        self.errors.append(f"Ligne {row}: {message}")
        self.stats['errors'] += 1


class TauxEmprunteurImporter(ExcelImporter):
    """
    Importer pour les taux emprunteur
    Format Excel attendu:
    | Tranche Age | Age Min | Age Max | Durée (années) | Taux (%) | Produit | Date Début |
    """
    
    def import_data(self, sheet_name='Taux', start_row=2, banque=None):
        """Importe les données depuis une feuille Excel"""
        
        if not self.load_workbook():
            return False
        
        try:
            ws = self.workbook[sheet_name]
        except KeyError:
            self.errors.append(f"Feuille '{sheet_name}' introuvable")
            return False
        
        with transaction.atomic():
            for row_num, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
                if not any(row):  # Ligne vide
                    continue
                
                try:
                    # Extraire les données
                    tranche_age = str(row[0]).strip() if row[0] else None
                    age_min = int(row[1]) if row[1] else None
                    age_max = int(row[2]) if row[2] else None
                    duree = int(row[3]) if row[3] else None
                    taux = Decimal(str(row[4])) if row[4] else None
                    produit = str(row[5]).strip() if row[5] and len(row) > 5 else 'emprunteur'
                    date_debut = row[6] if len(row) > 6 and row[6] else date.today()
                    
                    # Validation
                    if not all([tranche_age, age_min, age_max, duree, taux]):
                        self.log_error(row_num, "Données manquantes")
                        continue
                    
                    # Convertir date si nécessaire
                    if isinstance(date_debut, datetime):
                        date_debut = date_debut.date()
                    elif not isinstance(date_debut, date):
                        date_debut = date.today()
                    
                    # Créer ou mettre à jour
                    obj, created = TableTauxEmprunteur.objects.update_or_create(
                        age_min=age_min,
                        age_max=age_max,
                        duree_annees=duree,
                        produit=produit,
                        banque=banque,
                        defaults={
                            'tranche_age': tranche_age,
                            'taux_pourcentage': taux,
                            'date_debut_validite': date_debut,
                            'actif': True
                        }
                    )
                    
                    if created:
                        self.stats['created'] += 1
                    else:
                        self.stats['updated'] += 1
                
                except Exception as e:
                    self.log_error(row_num, str(e))
        
        return self.stats['errors'] == 0


class CIMA_H_Importer(ExcelImporter):
    """
    Importer pour la table CIMA H
    Format Excel attendu:
    | x (âge) | Nx | Mx | Dx | lx | dx | qx | Cx |
    """
    
    def import_data(self, sheet_name='CIMA_H', start_row=2):
        """Importe les données depuis une feuille Excel"""
        
        if not self.load_workbook():
            return False
        
        try:
            ws = self.workbook[sheet_name]
        except KeyError:
            self.errors.append(f"Feuille '{sheet_name}' introuvable")
            return False
        
        with transaction.atomic():
            # Supprimer les anciennes données (table de référence)
            TableCIMA_H.objects.all().delete()
            
            for row_num, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
                if not any(row):
                    continue
                
                try:
                    x = int(row[0]) if row[0] is not None else None
                    Nx = Decimal(str(row[1])) if row[1] is not None else None
                    Mx = Decimal(str(row[2])) if row[2] is not None else None
                    Dx = Decimal(str(row[3])) if row[3] is not None else None
                    lx = Decimal(str(row[4])) if row[4] is not None else None
                    dx = Decimal(str(row[5])) if row[5] is not None else None
                    qx = Decimal(str(row[6])) if row[6] is not None else None
                    Cx = Decimal(str(row[7])) if row[7] is not None else None
                    
                    if x is None or any(v is None for v in [Nx, Mx, Dx, lx, dx, qx, Cx]):
                        self.log_error(row_num, "Données manquantes")
                        continue
                    
                    TableCIMA_H.objects.create(
                        x=x, Nx=Nx, Mx=Mx, Dx=Dx,
                        lx=lx, dx=dx, qx=qx, Cx=Cx
                    )
                    self.stats['created'] += 1
                
                except Exception as e:
                    self.log_error(row_num, str(e))
        
        return self.stats['errors'] == 0


class PrimesEtudesImporter(ExcelImporter):
    """
    Importer pour les primes études
    Format Excel attendu:
    | Age | Durée Paiement | Durée Rente | Type Prime | Produit | Montant |
    """
    
    def import_data(self, sheet_name='Primes_Etudes', start_row=2):
        """Importe les données depuis une feuille Excel"""
        
        if not self.load_workbook():
            return False
        
        try:
            ws = self.workbook[sheet_name]
        except KeyError:
            self.errors.append(f"Feuille '{sheet_name}' introuvable")
            return False
        
        with transaction.atomic():
            for row_num, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
                if not any(row):
                    continue
                
                try:
                    age = int(row[0]) if row[0] else None
                    duree_paiement = int(row[1]) if row[1] else None
                    duree_rente = int(row[2]) if row[2] else None
                    type_prime = str(row[3]).strip().upper() if row[3] else None
                    produit = str(row[4]).strip() if row[4] else None
                    montant = Decimal(str(row[5])) if row[5] else None
                    
                    if not all([age, duree_paiement, duree_rente, type_prime, produit, montant]):
                        self.log_error(row_num, "Données manquantes")
                        continue
                    
                    obj, created = TablePrimesEtudes.objects.update_or_create(
                        age=age,
                        duree_paiement=duree_paiement,
                        duree_rente=duree_rente,
                        type_prime=type_prime,
                        produit=produit,
                        defaults={
                            'montant': montant,
                            'actif': True
                        }
                    )
                    
                    if created:
                        self.stats['created'] += 1
                    else:
                        self.stats['updated'] += 1
                
                except Exception as e:
                    self.log_error(row_num, str(e))
        
        return self.stats['errors'] == 0


class TauxMensuelsImporter(ExcelImporter):
    """
    Importer pour les taux mensuels
    Format Excel attendu:
    | Age | Durée Paiement | Durée Rente | Produit | Taux |
    """
    
    def import_data(self, sheet_name='Taux_Mensuels', start_row=2):
        """Importe les données depuis une feuille Excel"""
        
        if not self.load_workbook():
            return False
        
        try:
            ws = self.workbook[sheet_name]
        except KeyError:
            self.errors.append(f"Feuille '{sheet_name}' introuvable")
            return False
        
        with transaction.atomic():
            for row_num, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
                if not any(row):
                    continue
                
                try:
                    age = int(row[0]) if row[0] else None
                    duree_paiement = int(row[1]) if row[1] else None
                    duree_rente = int(row[2]) if row[2] else None
                    produit = str(row[3]).strip() if row[3] else 'NSIA-ETUDES'
                    taux = Decimal(str(row[4])) if row[4] else None
                    
                    if not all([age, duree_paiement, duree_rente, taux]):
                        self.log_error(row_num, "Données manquantes")
                        continue
                    
                    obj, created = TableTauxMensuels.objects.update_or_create(
                        age=age,
                        duree_paiement=duree_paiement,
                        duree_rente=duree_rente,
                        produit=produit,
                        defaults={
                            'taux': taux,
                            'actif': True
                        }
                    )
                    
                    if created:
                        self.stats['created'] += 1
                    else:
                        self.stats['updated'] += 1
                
                except Exception as e:
                    self.log_error(row_num, str(e))
        
        return self.stats['errors'] == 0
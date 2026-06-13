# apps/simulateur/services/export_service.py
"""
Service d'export des simulations en CSV et JSON
Gère le mapping intelligent des données selon le produit
"""

import csv
import json
from io import StringIO, BytesIO
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import List, Dict, Any, Optional
from uuid import UUID
from django.db.models import QuerySet
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side



class ExportService:
    """
    Service pour exporter les simulations en CSV ou JSON
    Supporte tous les produits : emprunteur, retraite, etudes, elikia, mobateli
    """
    
    # Les 21 colonnes CSV fixes pour tous les produits
    COLONNES_CSV = [
        "Numéro",
        "Contrat",
        "Produit",
        "N° de Proposition",
        "Durée",
        "Prime Nette Hors Accessoires",
        "Accessoires",
        "Prime Périodique",
        "Périodicité Cotisation",
        "Capital",
        "Date d'Émission",
        "Date d'Effet",
        "Date d'Échéance",
        "Souscripteur",
        "Compte Souscripteur",
        "Assuré",
        "Date de Naissance Assuré",
        "Numéro Téléphone",
        "Compte Assuré",
        "Statut",
        "Nom du Vendeur",
    ]
    
    def __init__(self):
        """Initialise le service d'export"""
        pass
    
    def exporter_csv(self, simulations: QuerySet, filtres: Dict = None) -> BytesIO:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Simulations"

        header_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # EN-TÊTES
        row_index = 1
        for col_index, colonne in enumerate(self.COLONNES_CSV, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=colonne)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            sheet.column_dimensions[cell.column_letter].width = 25

        # DONNÉES
        row_index = 2
        for simulation in simulations:
            ligne = self._extraire_ligne_csv(simulation)

            for col_index, colonne in enumerate(self.COLONNES_CSV, start=1):
                valeur = ligne.get(colonne, "")
                valeur_excel = self.normaliser_valeur_excel(valeur)

                cell = sheet.cell(row=row_index, column=col_index, value=valeur_excel)
                cell.border = border

            row_index += 1

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def exporter_json(self, simulations: QuerySet, filtres: Dict = None) -> BytesIO:
        """
        Exporte les simulations en JSON complet
        Inclut donnees_entree + resultats_calcul + métadonnées
        
        Args:
            simulations: QuerySet des simulations à exporter
            filtres: Dictionnaire des filtres appliqués
            
        Returns:
            BytesIO: Contenu du fichier JSON
        """
        data = []
        
        for simulation in simulations:
            data.append({
                "id": str(simulation.id),
                "reference": simulation.reference,
                "banque": {
                    "id": str(simulation.banque.id),
                    "nom": simulation.banque.nom,
                    "code": simulation.banque.code_banque,
                },
                "gestionnaire": {
                    "id": str(simulation.gestionnaire.id) if simulation.gestionnaire else None,
                    "nom": simulation.gestionnaire.get_full_name() if simulation.gestionnaire else None,
                },
                "produit": simulation.produit,
                "produit_display": simulation.get_produit_display(),
                "statut": simulation.statut,
                "statut_display": simulation.get_statut_display(),
                
                # Client
                "nom_client": simulation.nom_client,
                "prenom_client": simulation.prenom_client,
                "email_client": simulation.email_client,
                "telephone_client": simulation.telephone_client,
                
                # Données complètes
                "donnees_entree": simulation.donnees_entree,
                "resultats_calcul": simulation.resultats_calcul,
                
                # Métadonnées
                "date_creation": simulation.date_creation.isoformat(),
                "date_modification": simulation.date_modification.isoformat(),
                "date_validation": simulation.date_validation.isoformat() if simulation.date_validation else None,
                "notes": simulation.notes or "",
            })
        
        # Convertir en BytesIO
        json_str = json.dumps(data, indent=2, ensure_ascii=False, default=str)
        json_bytes = BytesIO(json_str.encode('utf-8'))
        json_bytes.seek(0)
        
        return json_bytes
    
    def _extraire_ligne_csv(self, simulation) -> Dict[str, Any]:
        """
        Extrait une ligne CSV à partir d'une simulation
        Mappe intelligemment selon le produit
        
        Args:
            simulation: Instance de Simulation
            
        Returns:
            dict: Ligne CSV avec les 21 colonnes
        """
        produit = simulation.produit
        donnees = simulation.donnees_entree or {}
        resultats = simulation.resultats_calcul or {}
        
        # Mapping selon le produit
        if produit == 'emprunteur':
            return self._mapper_emprunteur(simulation, donnees, resultats)
        elif produit == 'retraite':
            return self._mapper_retraite(simulation, donnees, resultats)
        elif produit == 'etudes':
            return self._mapper_etudes(simulation, donnees, resultats)
        elif produit == 'elikia':
            return self._mapper_elikia(simulation, donnees, resultats)
        elif produit == 'mobateli':
            return self._mapper_mobateli(simulation, donnees, resultats)
        elif produit == 'epargne_plus':
            return self._mapper_epargne_plus(simulation, donnees, resultats)
        else:
            # Produit inconnu : mapper les champs communs
            #return self._mapper_generique(simulation, donnees, resultats)
            return None
    
    def _mapper_emprunteur(self, simulation, donnees: Dict, resultats: Dict) -> Dict:
        """Mapping spécifique pour Assurance Emprunteur"""
        
        return {
            "Numéro": simulation.id,
            "Contrat": getattr(simulation, 'numero_contrat', '') or '',
            "Produit": "EMPRUNTEUR",
            "N° de Proposition": simulation.reference,
            "Durée": resultats.get('duree_mois', ''),
            "Prime Nette Hors Accessoires": self._format_decimal(resultats.get('prime_nette', 0)),
            "Accessoires": self._format_decimal(resultats.get('frais_accessoires', 0)),
            "Prime Périodique": self._calculer_prime_periodique(
                resultats.get('prime_totale', 0),
                resultats.get('duree_mois', 1)
            ),
            "Périodicité Cotisation": "Mensuelle",
            "Capital": self._format_decimal(resultats.get('montant_pret', 0)),
            "Date d'Émission": self._format_date(simulation.date_creation),
            "Date d'Effet": self._format_date(resultats.get('date_effet', '')),
            "Date d'Échéance": self._format_date(resultats.get('date_terme', '')),
            "Souscripteur": f"{simulation.nom_client} {simulation.prenom_client}",
            "Compte Souscripteur": donnees.get('numero_compte', ''),
            "Assuré": f"{simulation.nom_client} {simulation.prenom_client}",
            "Date de Naissance Assuré": self._format_date(donnees.get('date_naissance', '')),
            "Numéro Téléphone": simulation.telephone_client or '',
            "Compte Assuré": donnees.get('numero_compte', ''),
            "Statut": simulation.get_statut_display(),
            "Nom du Vendeur": self._get_gestionnaire_nom(simulation),

        }
    
    def _mapper_retraite(self, simulation, donnees: Dict, resultats: Dict) -> Dict:
        """Mapping spécifique pour Confort Retraite"""
        prime_periodique = resultats.get('prime_periodique_commerciale', 0)
        
        return {
            "Numéro": simulation.id,
            "Contrat": getattr(simulation, 'numero_contrat', '') or '',
            "Produit": "CONFORT RETRAITE",
            "N° de Proposition": simulation.reference,
            "Durée": f"{donnees.get('duree_contrat', '')} ans",
            "Prime Nette Hors Accessoires": self._format_decimal(prime_periodique - 1000),
            "Accessoires": "1000",
            "Prime Périodique": self._format_decimal(prime_periodique),
            "Périodicité Cotisation": donnees.get('periodicite', '').capitalize(),
            "Capital": self._format_decimal(resultats.get('capital_garanti', 0)),
            "Date d'Émission": self._format_date(simulation.date_creation),
            "Date d'Effet": self._format_date(donnees.get('date_effet', '')),
            "Date d'Échéance": self._calculer_date_echeance_retraite(
                donnees.get('date_effet', ''),
                donnees.get('duree_contrat', 0)
            ),
            "Souscripteur": f"{simulation.nom_client} {simulation.prenom_client}",
            "Compte Souscripteur": donnees.get('numero_compte', ''),
            "Assuré": f"{simulation.nom_client} {simulation.prenom_client}",
            "Date de Naissance Assuré": self._format_date(donnees.get('date_naissance', '')),
            "Numéro Téléphone": simulation.telephone_client or '',
            "Compte Assuré": donnees.get('numero_compte', ''),
            "Statut": simulation.get_statut_display(),
            "Nom du Vendeur": self._get_gestionnaire_nom(simulation),
        }
    
    def _mapper_etudes(self, simulation, donnees: Dict, resultats: Dict) -> Dict:
        """Mapping spécifique pour Confort Études"""
        duree_totale = donnees.get('duree_cotisation', 0) + donnees.get('duree_service_rente', 0)
        
        return {
            "Numéro": simulation.id,
            "Contrat": getattr(simulation, 'numero_contrat', '') or '',
            "Produit": "CONFORT ETUDES",
            "N° de Proposition": simulation.reference,
            "Durée": f"{duree_totale} ans",
            "Prime Nette Hors Accessoires": self._format_decimal(resultats.get('prime_annuelle', 0)),
            "Accessoires": "1000",
            "Prime Périodique": self._format_decimal(resultats.get('prime_mensuelle', 0)),
            "Périodicité Cotisation": donnees.get('periodicite', '').capitalize(),
            "Capital": self._format_decimal(resultats.get('montant_rente_annuel', 0)),
            "Date d'Émission": self._format_date(simulation.date_creation),
            "Date d'Effet": self._format_date(donnees.get('date_effet', '')),
            "Date d'Échéance": self._calculer_date_echeance_etudes(
                donnees.get('date_effet', ''),
                duree_totale
            ),
            "Souscripteur": f"{simulation.nom_client} {simulation.prenom_client}",
            "Compte Souscripteur": donnees.get('numero_compte', ''),
            "Assuré": self._get_beneficiaire_principal(simulation),
            "Date de Naissance Assuré": self._format_date(donnees.get('date_naissance_enfant', '')),
            "Numéro Téléphone": simulation.telephone_client or '',
            "Compte Assuré": donnees.get('numero_compte', ''),
            "Statut": simulation.get_statut_display(),
            "Nom du Vendeur": self._get_gestionnaire_nom(simulation),
        }
    
    def _mapper_elikia(self, simulation, donnees: Dict, resultats: Dict) -> Dict:
        """Mapping spécifique pour Elikia Scolaire"""
        return {
            "Numéro": simulation.id,
            "Contrat": getattr(simulation, 'numero_contrat', '') or '',
            "Produit": "ELIKIA SCOLAIRE",
            "N° de Proposition": simulation.reference,
            "Durée": f"{donnees.get('duree_rente', '')} ans",
            "Prime Nette Hors Accessoires": self._format_decimal(resultats.get('prime_nette_annuelle', 0)),
            "Accessoires": self._format_decimal(resultats.get('frais_accessoires', 0)),
            "Prime Périodique": self._format_decimal(resultats.get('prime_totale', 0)),
            "Périodicité Cotisation": donnees.get('periodicite'),
            "Capital": self._format_decimal(resultats.get('capital_garanti', donnees.get('capital_dtc_iad', 0))),
            "Date d'Émission": self._format_date(simulation.date_creation),
            "Date d'Effet": self._format_date(donnees.get('date_effet', '')),
            "Date d'Échéance": self._calculer_date_echeance_elikia(
                donnees.get('date_effet', ''),
                donnees.get('duree_rente', 0)
            ),
            "Souscripteur": f"{simulation.nom_client} {simulation.prenom_client}",
            "Compte Souscripteur": donnees.get('numero_compte', ''),
            "Assuré": f"{simulation.nom_client} {simulation.prenom_client}",
            "Date de Naissance Assuré": self._format_date(donnees.get('date_naissance', '')),
            "Numéro Téléphone": simulation.telephone_client or '',
            "Compte Assuré": donnees.get('numero_compte', ''),
            "Statut": simulation.get_statut_display(),
            "Nom du Vendeur": self._get_gestionnaire_nom(simulation),
        }
    
    def _mapper_mobateli(self, simulation, donnees: Dict, resultats: Dict) -> Dict:
        """Mapping spécifique pour Mobateli (DTC/IAD)"""
        return {
            "Numéro": simulation.id,
            "Contrat": getattr(simulation, 'numero_contrat', '') or '',
            "Produit": "MOBATELI",
            "N° de Proposition": simulation.reference,
            "Durée": "",  # Pas de durée pour Mobateli
            "Prime Nette Hors Accessoires": self._format_decimal(resultats.get('prime_nette', 0)),
            "Accessoires": self._format_decimal(resultats.get('frais_accessoires', 0)),
            "Prime Périodique": self._format_decimal(resultats.get('prime_totale', 0)),
            "Périodicité Cotisation": donnees.get('periodicite'),
            "Capital": self._format_decimal(donnees.get('capital_dtc_iad', 0)),
            "Date d'Émission": self._format_date(simulation.date_creation),
            "Date d'Effet": self._format_date(donnees.get('date_effet', '')),
            "Date d'Échéance": "",  # Pas d'échéance
            "Souscripteur": f"{simulation.nom_client} {simulation.prenom_client}",
            "Compte Souscripteur": donnees.get('numero_compte', ''),
            "Assuré": f"{simulation.nom_client} {simulation.prenom_client}",
            "Date de Naissance Assuré": self._format_date(donnees.get('date_naissance', '')),
            "Numéro Téléphone": simulation.telephone_client or '',
            "Compte Assuré": donnees.get('numero_compte', ''),
            "Statut": simulation.get_statut_display(),
            "Nom du Vendeur": self._get_gestionnaire_nom(simulation),
        }

    def _mapper_epargne_plus(self, simulation, donnees: Dict, resultats: Dict) -> Dict:
        """Mapping Épargne Plus – structure commune obligatoire"""

        return {
            "Numéro": simulation.id,
            "Contrat": getattr(simulation, 'numero_contrat', '') or '',
            "Produit": simulation.get_produit_display(),
            "N° de Proposition": simulation.reference,

            # Champs communs mais adaptés à l’épargne
            "Durée": donnees.get('duree_annees', ''),
            "Prime Nette Hors Accessoires": self._format_decimal(
                donnees.get('cotisation_mensuelle', 0)
            ),
            "Accessoires": self._format_decimal(
                resultats.get('frais_adhesion', 0)
            ),
            "Prime Périodique": self._format_decimal(
                donnees.get('cotisation_mensuelle', 0) + resultats.get('frais_adhesion', 0)
            ),
            "Périodicité Cotisation": donnees.get('periodicite', ''),
            "Capital": self._format_decimal(
                resultats.get('capital_acquis', 0)
            ),

            # Dates
            "Date d'Émission": self._format_date(simulation.date_creation),
            "Date d'Effet": self._format_date(
                donnees.get('date_effet', '')
            ),
            "Date d'Échéance": self._format_date(
                resultats.get('date_terme', '')
            ),

            # Personnes
            "Souscripteur": f"{simulation.nom_client} {simulation.prenom_client}",
            "Compte Souscripteur": donnees.get('numero_compte', ''),
            "Assuré": f"{simulation.nom_client} {simulation.prenom_client}",
            "Date de Naissance Assuré": self._format_date(
                donnees.get('date_naissance', '')
            ),

            # Contact & gestion
            "Numéro Téléphone": simulation.telephone_client or '',
            "Compte Assuré": donnees.get('numero_compte', ''),
            "Statut": simulation.get_statut_display(),
            "Nom du Vendeur": self._get_gestionnaire_nom(simulation),
        }

    def _mapper_generique(self, simulation, donnees: Dict, resultats: Dict) -> Dict:
        """Mapping générique pour produits inconnus"""
        return {
            "Numéro": simulation.id,
            "Contrat": getattr(simulation, 'numero_contrat', '') or '',
            "Produit": simulation.get_produit_display(),
            "N° de Proposition": simulation.reference,
            "Durée": "",
            "Prime Nette Hors Accessoires": "",
            "Accessoires": "",
            "Prime Périodique": "",
            "Périodicité Cotisation": "",
            "Capital": "",
            "Date d'Émission": self._format_date(simulation.date_creation),
            "Date d'Effet": "",
            "Date d'Échéance": "",
            "Souscripteur": f"{simulation.nom_client} {simulation.prenom_client}",
            "Compte Souscripteur": "",
            "Assuré": f"{simulation.nom_client} {simulation.prenom_client}",
            "Date de Naissance Assuré": "",
            "Numéro Téléphone": simulation.telephone_client or '',
            "Compte Assuré": "",
            "Statut": simulation.get_statut_display(),
            "Nom du Vendeur": self._get_gestionnaire_nom(simulation),
        }
    
    # ============================================
    # MÉTHODES UTILITAIRES
    # ============================================

    def normaliser_valeur_excel(self, valeur):
            """
            Convertit les types Python non supportés par Excel
            """
            if valeur is None:
                return ""

            if isinstance(valeur, UUID):
                return str(valeur)

            if isinstance(valeur, Decimal):
                return float(valeur)

            if isinstance(valeur, (datetime, date)):
                return valeur

            return valeur
    
    def _format_decimal(self, valeur) -> str:
        """Formate un Decimal ou float en string"""
        if valeur is None or valeur == '':
            return ''
        try:
            return f"{float(valeur):.2f}"
        except (ValueError, TypeError):
            return ''
    
    def _format_date(self, date) -> str:
        """Formate une date en string DD/MM/YYYY"""
        if not date:
            return ''
        
        if isinstance(date, str):
            # Essayer de parser si c'est une string
            try:
                if 'T' in date:  # Format ISO
                    date = datetime.fromisoformat(date.replace('Z', '+00:00'))
                else:
                    # Essayer format YYYY-MM-DD
                    date = datetime.strptime(date, '%Y-%m-%d')
            except (ValueError, AttributeError):
                return date  # Retourner tel quel si pas parsable
        
        if isinstance(date, (datetime, timezone.datetime)):
            return date.strftime('%d/%m/%Y')
        
        return str(date)
    
    def _calculer_prime_periodique(self, prime_totale, duree_mois) -> str:
        """Calcule la prime mensuelle"""
        if not prime_totale or not duree_mois or duree_mois == 0:
            return ''
        try:
            prime_mensuelle = float(prime_totale) / float(duree_mois)
            return f"{prime_mensuelle:.2f}"
        except (ValueError, TypeError, ZeroDivisionError):
            return ''
    
    def _calculer_date_echeance_retraite(self, date_effet, duree_annees) -> str:
        """Calcule la date d'échéance pour Retraite"""
        if not date_effet or not duree_annees:
            return ''
        
        try:
            if isinstance(date_effet, str):
                date_debut = datetime.strptime(date_effet, '%Y-%m-%d')
            else:
                date_debut = date_effet
            
            date_echeance = date_debut + timedelta(days=365 * int(duree_annees))
            return date_echeance.strftime('%d/%m/%Y')
        except (ValueError, TypeError, AttributeError):
            return ''
    
    def _calculer_date_echeance_etudes(self, date_effet, duree_totale_annees) -> str:
        """Calcule la date d'échéance pour Études"""
        return self._calculer_date_echeance_retraite(date_effet, duree_totale_annees)
    
    def _calculer_date_echeance_elikia(self, date_effet, duree_service_annees) -> str:
        """Calcule la date d'échéance pour Elikia"""
        return self._calculer_date_echeance_retraite(date_effet, duree_service_annees)
    
    def _get_gestionnaire_nom(self, simulation) -> str:
        """Récupère le nom complet du gestionnaire"""
        if hasattr(simulation, 'gestionnaire') and simulation.gestionnaire:
            return simulation.gestionnaire.get_full_name()
        elif hasattr(simulation, 'gestionnaire_nom'):
            return simulation.gestionnaire_nom
        return ''
    
    def _get_beneficiaire_principal(self, simulation) -> str:
        """
        Récupère le nom du bénéficiaire principal (pour Études)
        Généralement l'enfant
        """
        # Essayer de récupérer depuis les bénéficiaires
        try:
            beneficiaires = getattr(simulation, 'beneficiaires', None)
            if beneficiaires and beneficiaires.exists():
                principal = beneficiaires.first()
                return principal.nom_prenoms
        except:
            pass
        
        # Sinon, retourner le souscripteur
        return f"{simulation.nom_client} {simulation.prenom_client}"
    
    def generer_nom_fichier(self, format_export: str, filtres: Dict = None) -> str:
        """
        Génère le nom de fichier selon les filtres appliqués
        Format: Export_Simulations_BANQUE_AGENCE_AAAAMMJJ.csv
        
        Args:
            format_export: 'csv' ou 'json'
            filtres: Dictionnaire des filtres appliqués
            
        Returns:
            str: Nom du fichier
        """
        parts = ["Export", "Simulations"]
        
        if filtres:
            if filtres.get('banque'):
                parts.append(filtres['banque'].upper())
            if filtres.get('agence'):
                parts.append(filtres['agence'].replace(' ', '_'))
            if filtres.get('gestionnaire'):
                parts.append(f"Gest{filtres['gestionnaire']}")
            if filtres.get('produit'):
                parts.append(filtres['produit'].upper())
        
        # Ajouter la date
        date_str = datetime.now().strftime('%Y%m%d')
        parts.append(date_str)
        
        # Construire le nom
        nom = '_'.join(parts)
        extension = 'xlsx' if format_export == 'xlsx' else 'json'
        
        return f"{nom}.{extension}"


# ============================================
# FONCTION HELPER
# ============================================

def exporter_simulations(simulations: QuerySet, format_export: str = 'csv', filtres: Dict = None):
    """
    Fonction helper pour exporter rapidement
    
    Args:
        simulations: QuerySet de simulations
        format_export: 'csv' ou 'json'
        filtres: Dictionnaire des filtres
        
    Returns:
        tuple: (BytesIO du fichier, nom du fichier)
    """
    service = ExportService()
    
    if format_export == 'csv':
        fichier = service.exporter_csv(simulations, filtres)
    else:
        fichier = service.exporter_json(simulations, filtres)
    
    nom_fichier = service.generer_nom_fichier(format_export, filtres)
    
    return fichier, nom_fichier